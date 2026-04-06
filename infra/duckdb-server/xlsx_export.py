from datetime import datetime, timezone
from io import BytesIO

from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from source_links import DATASET_URLS, build_bbl_urls, build_source_url

# ── Colours ────────────────────────────────────────────────────────────────────
_DARK_BLUE = "1A2744"
_MED_BLUE = "2C3E6B"
_LIGHT_BLUE = "E8EDF5"
_LINK_BLUE = "1155CC"
_BORDER_GREY = "DDDDDD"

_FILL_DARK = PatternFill("solid", fgColor=_DARK_BLUE)
_FILL_MED = PatternFill("solid", fgColor=_MED_BLUE)
_FILL_LIGHT = PatternFill("solid", fgColor=_LIGHT_BLUE)

_FONT_TITLE = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
_FONT_META = Font(name="Calibri", size=10, color="888888")
_FONT_HEADER = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
_FONT_DATA = Font(name="Calibri", size=11)
_FONT_LINK = Font(name="Calibri", size=11, color=_LINK_BLUE, underline="single")
_FONT_FOOTER = Font(name="Calibri", size=9, color="888888", italic=True)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")

_THIN_BOTTOM = Border(bottom=Side(style="thin", color=_BORDER_GREY))


def _col_width(name: str) -> float:
    name_lower = name.lower()
    if name == "#":
        return 5
    if name_lower == "bbl":
        return 14
    if name_lower in ("name", "address") or "name" in name_lower or "address" in name_lower:
        return 30
    return min(max(len(name) + 2, 12), 40)


def generate_branded_xlsx(
    cols: list,
    rows: list,
    table_name: str = "",
    tool_name: str = "",
    sql: str = "",
    sources=None,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = table_name or "Export"

    n_rows = len(rows)
    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    sources_str = ", ".join(sources) if sources else "NYC Open Data"

    # Determine whether to add Source column
    has_source_col = bool(table_name and table_name in DATASET_URLS)

    # Total columns: # + data cols + optional Source
    n_cols = 1 + len(cols) + (1 if has_source_col else 0)
    last_col = get_column_letter(n_cols)

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col}1")
    title_cell = ws["A1"]
    title_cell.value = "COMMON GROUND — NYC Open Data Export"
    title_cell.fill = _FILL_DARK
    title_cell.font = _FONT_TITLE
    title_cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[1].height = 24

    # ── Row 2: Metadata ───────────────────────────────────────────────────────
    ws.merge_cells(f"A2:{last_col}2")
    meta_cell = ws["A2"]
    meta_cell.value = (
        f"common-ground.nyc | Generated {timestamp} | {n_rows} rows | Sources: {sources_str}"
    )
    meta_cell.fill = _FILL_LIGHT
    meta_cell.font = _FONT_META
    meta_cell.alignment = _ALIGN_LEFT
    ws.row_dimensions[2].height = 18

    # ── Row 3: Headers ────────────────────────────────────────────────────────
    headers = ["#"] + list(cols) + (["Source"] if has_source_col else [])
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = _FILL_MED
        cell.font = _FONT_HEADER
        cell.alignment = _ALIGN_CENTER
    ws.row_dimensions[3].height = 20

    # ── Freeze panes ──────────────────────────────────────────────────────────
    ws.freeze_panes = "A4"

    # ── Column widths ─────────────────────────────────────────────────────────
    for col_idx, header in enumerate(headers, start=1):
        if header == "Source":
            ws.column_dimensions[get_column_letter(col_idx)].width = 10
        else:
            ws.column_dimensions[get_column_letter(col_idx)].width = _col_width(header)

    # ── Identify special columns ──────────────────────────────────────────────
    bbl_col_indices = [
        i + 2 for i, c in enumerate(cols) if str(c).lower() == "bbl"
    ]  # 1-based, offset by 2 (# col is 1)

    percentile_col_indices = [
        i + 2
        for i, c in enumerate(cols)
        if "percentile" in str(c).lower() or "pctile" in str(c).lower()
    ]

    source_col_idx = n_cols if has_source_col else None

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_num, row_data in enumerate(rows, start=1):
        sheet_row = row_num + 3  # rows 1-3 are title/meta/header
        row_dict = {col: val for col, val in zip(cols, row_data)}

        # Row ID
        id_cell = ws.cell(row=sheet_row, column=1, value=row_num)
        id_cell.font = _FONT_DATA
        id_cell.border = _THIN_BOTTOM
        id_cell.alignment = _ALIGN_CENTER

        # Data cells
        for col_idx, (col, val) in enumerate(zip(cols, row_data), start=2):
            cell = ws.cell(row=sheet_row, column=col_idx, value=val)
            cell.border = _THIN_BOTTOM

            if col_idx in bbl_col_indices and val:
                urls = build_bbl_urls(str(val))
                if urls:
                    cell.hyperlink = urls[0]["url"]
                    cell.font = _FONT_LINK
                else:
                    cell.font = _FONT_DATA
            else:
                cell.font = _FONT_DATA

        # Source column
        if source_col_idx is not None:
            src_url = build_source_url(table_name, row_dict)
            src_cell = ws.cell(row=sheet_row, column=source_col_idx, value="View →")
            src_cell.border = _THIN_BOTTOM
            if src_url:
                src_cell.hyperlink = src_url
                src_cell.font = _FONT_LINK
            else:
                src_cell.font = _FONT_DATA

    # ── Percentile conditional formatting ────────────────────────────────────
    if rows:
        data_start = 4
        data_end = 3 + n_rows
        for col_idx in percentile_col_indices:
            col_letter = get_column_letter(col_idx)
            rule = ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="63BE7B",  # green
                mid_type="num",
                mid_value=50,
                mid_color="FFEB84",  # yellow
                end_type="num",
                end_value=100,
                end_color="F8696B",  # red
            )
            ws.conditional_formatting.add(
                f"{col_letter}{data_start}:{col_letter}{data_end}", rule
            )

    # ── Footer ────────────────────────────────────────────────────────────────
    footer_row = 3 + n_rows + 2
    ws.merge_cells(f"A{footer_row}:{last_col}{footer_row}")
    footer_cell = ws[f"A{footer_row}"]
    footer_cell.value = (
        "Data: NYC public records (open data, no restrictions) | "
        "Powered by DuckDB + DuckLake | common-ground.nyc"
    )
    footer_cell.font = _FONT_FOOTER
    footer_cell.alignment = _ALIGN_LEFT

    # ── Serialise ─────────────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
