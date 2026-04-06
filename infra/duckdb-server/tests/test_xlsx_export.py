import os
import sys
from io import BytesIO

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from openpyxl import load_workbook

from xlsx_export import generate_branded_xlsx


COLS = ["name", "address", "bbl"]
ROWS = [
    ("Alice", "123 Main St", "1000670001"),
    ("Bob", "456 Elm Ave", "2005230010"),
]


def _make_wb(cols, rows, **kwargs):
    buf = generate_branded_xlsx(cols, rows, **kwargs)
    return load_workbook(BytesIO(buf))


# --- Return type ---

def test_returns_bytes():
    result = generate_branded_xlsx(COLS, ROWS)
    assert isinstance(result, bytes)


def test_returns_non_empty():
    result = generate_branded_xlsx(COLS, ROWS)
    assert len(result) > 0


# --- Row 1: title ---

def test_row1_contains_common_ground():
    wb = _make_wb(COLS, ROWS)
    ws = wb.active
    val = ws.cell(row=1, column=1).value
    assert val is not None
    assert "COMMON GROUND" in str(val)


# --- Row 2: metadata ---

def test_row2_contains_generated():
    wb = _make_wb(COLS, ROWS)
    ws = wb.active
    val = ws.cell(row=2, column=1).value
    assert val is not None
    assert "Generated" in str(val)


def test_row2_contains_row_count():
    wb = _make_wb(COLS, ROWS)
    ws = wb.active
    val = ws.cell(row=2, column=1).value
    assert "2" in str(val)


def test_row2_contains_sources_when_provided():
    wb = _make_wb(COLS, ROWS, sources=["HPD violations"])
    ws = wb.active
    val = ws.cell(row=2, column=1).value
    assert "HPD violations" in str(val)


def test_row2_zero_rows():
    wb = _make_wb(COLS, [], sources=None)
    ws = wb.active
    val = ws.cell(row=2, column=1).value
    assert "0" in str(val)


# --- Row 3: headers ---

def test_row3_first_col_is_hash():
    wb = _make_wb(COLS, ROWS)
    ws = wb.active
    val = ws.cell(row=3, column=1).value
    assert val == "#"


def test_row3_headers_match_cols():
    wb = _make_wb(COLS, ROWS)
    ws = wb.active
    for i, col in enumerate(COLS, start=2):
        assert ws.cell(row=3, column=i).value == col


# --- Row 4+: data ---

def test_row4_col1_is_1():
    wb = _make_wb(COLS, ROWS)
    ws = wb.active
    assert ws.cell(row=4, column=1).value == 1


def test_row5_col1_is_2():
    wb = _make_wb(COLS, ROWS)
    ws = wb.active
    assert ws.cell(row=5, column=1).value == 2


def test_row4_col2_is_first_data_value():
    wb = _make_wb(COLS, ROWS)
    ws = wb.active
    assert ws.cell(row=4, column=2).value == "Alice"


def test_row5_col2_is_second_data_value():
    wb = _make_wb(COLS, ROWS)
    ws = wb.active
    assert ws.cell(row=5, column=2).value == "Bob"


# --- Freeze panes ---

def test_freeze_panes_set():
    wb = _make_wb(COLS, ROWS)
    ws = wb.active
    assert ws.freeze_panes is not None


# --- Footer ---

def test_footer_contains_common_ground():
    wb = _make_wb(COLS, ROWS)
    ws = wb.active
    # Footer is last row + 2 after data, so row 4+2 rows of data + 2 = row 8
    # Scan last few rows for footer content
    found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "common-ground" in str(cell.value).lower():
                found = True
    assert found


def test_footer_contains_duckdb():
    wb = _make_wb(COLS, ROWS)
    ws = wb.active
    found = False
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and "DuckDB" in str(cell.value):
                found = True
    assert found


# --- BBL hyperlinks ---

def test_bbl_column_cells_have_hyperlinks():
    wb = _make_wb(COLS, ROWS)
    ws = wb.active
    # Find bbl column index (col 3 header = "bbl" → col 4 in sheet = col index 4)
    bbl_col = None
    for cell in ws[3]:
        if cell.value == "bbl":
            bbl_col = cell.column
            break
    assert bbl_col is not None, "bbl column not found in headers"
    # Check data rows have hyperlinks
    cell = ws.cell(row=4, column=bbl_col)
    assert cell.hyperlink is not None
    assert "cityofnewyork" in cell.hyperlink.target.lower() or "bisweb" in cell.hyperlink.target.lower()


# --- Source column ---

def test_source_column_added_when_table_name_known():
    wb = _make_wb(COLS, ROWS, table_name="hpd_violations")
    ws = wb.active
    # Last column in row 3 should be "Source"
    max_col = ws.max_column
    assert ws.cell(row=3, column=max_col).value == "Source"


def test_source_column_not_added_when_unknown_table():
    wb = _make_wb(COLS, ROWS, table_name="unknown_table_xyz")
    ws = wb.active
    max_col = ws.max_column
    # Last header should NOT be "Source"
    assert ws.cell(row=3, column=max_col).value != "Source"


def test_source_cells_have_hyperlinks():
    wb = _make_wb(COLS, ROWS, table_name="hpd_violations")
    ws = wb.active
    max_col = ws.max_column
    cell = ws.cell(row=4, column=max_col)
    assert cell.hyperlink is not None


# --- Percentile columns don't crash ---

def test_percentile_column_generates_without_crash():
    cols = ["name", "score_percentile"]
    rows = [("Alice", 75), ("Bob", 25)]
    buf = generate_branded_xlsx(cols, rows)
    assert isinstance(buf, bytes)
    assert len(buf) > 0


def test_pctile_column_generates_without_crash():
    cols = ["name", "rent_pctile"]
    rows = [("Alice", 80), ("Bob", 20)]
    buf = generate_branded_xlsx(cols, rows)
    assert isinstance(buf, bytes)
    assert len(buf) > 0


# --- Empty rows ---

def test_empty_rows_metadata_says_zero():
    wb = _make_wb(COLS, [])
    ws = wb.active
    val = str(ws.cell(row=2, column=1).value)
    assert "0" in val


def test_empty_rows_no_data_rows():
    wb = _make_wb(COLS, [])
    ws = wb.active
    # Row 4 should be footer or empty, not data
    val = ws.cell(row=4, column=1).value
    # Should not be row_id=1
    assert val != 1
